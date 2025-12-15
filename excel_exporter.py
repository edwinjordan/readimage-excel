"""
Excel Exporter Module
Exports extracted image features to Excel files.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Any
import os


class ExcelExporter:
    """Export image features to Excel format."""
    
    def __init__(self):
        """Initialize the Excel exporter."""
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self):
        """Create a new Excel workbook."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Image Features"
    
    def style_header(self):
        """Apply styling to the header row."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in self.worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    def auto_adjust_column_width(self):
        """Automatically adjust column widths based on content."""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except (TypeError, AttributeError):
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_single_image(self, features: Dict[str, Any], output_path: str):
        """
        Export features from a single image to Excel.
        
        Args:
            features: Dictionary containing extracted features
            output_path: Path where the Excel file will be saved
        """
        self.create_workbook()
        
        # Write headers and values
        self.worksheet.append(["Feature", "Value"])
        
        for key, value in features.items():
            self.worksheet.append([key, value])
        
        self.style_header()
        self.auto_adjust_column_width()
        
        # Save the workbook
        self.workbook.save(output_path)
        print(f"Excel file saved successfully: {output_path}")
    
    def export_multiple_images(self, features_list: List[Dict[str, Any]], output_path: str):
        """
        Export features from multiple images to Excel.
        
        Args:
            features_list: List of dictionaries containing extracted features
            output_path: Path where the Excel file will be saved
        """
        if not features_list:
            raise ValueError("No features to export")
        
        self.create_workbook()
        
        # Get all unique keys from all feature dictionaries
        all_keys = set()
        for features in features_list:
            all_keys.update(features.keys())
        
        headers = sorted(list(all_keys))
        
        # Write headers
        self.worksheet.append(headers)
        
        # Write data rows
        for features in features_list:
            row = [features.get(key, "") for key in headers]
            self.worksheet.append(row)
        
        self.style_header()
        self.auto_adjust_column_width()
        
        # Save the workbook
        self.workbook.save(output_path)
        print(f"Excel file saved successfully: {output_path}")
    
    def export_with_summary(self, features_list: List[Dict[str, Any]], output_path: str):
        """
        Export features with a summary sheet.
        
        Args:
            features_list: List of dictionaries containing extracted features
            output_path: Path where the Excel file will be saved
        """
        if not features_list:
            raise ValueError("No features to export")
        
        self.workbook = Workbook()
        
        # Create data sheet
        data_sheet = self.workbook.active
        data_sheet.title = "Image Data"
        
        # Get all unique keys
        all_keys = set()
        for features in features_list:
            all_keys.update(features.keys())
        headers = sorted(list(all_keys))
        
        # Write headers and data
        data_sheet.append(headers)
        for features in features_list:
            row = [features.get(key, "") for key in headers]
            data_sheet.append(row)
        
        # Style data sheet
        self.worksheet = data_sheet
        self.style_header()
        self.auto_adjust_column_width()
        
        # Create summary sheet
        summary_sheet = self.workbook.create_sheet("Summary", 0)
        summary_sheet.append(["Summary", ""])
        summary_sheet.append(["Total Images Processed", len(features_list)])
        
        # Calculate some statistics
        if 'avg_brightness' in headers:
            avg_brightness = sum(f.get('avg_brightness', 0) for f in features_list) / len(features_list)
            summary_sheet.append(["Average Brightness", round(avg_brightness, 2)])
        
        if 'width' in headers and 'height' in headers:
            avg_width = sum(f.get('width', 0) for f in features_list) / len(features_list)
            avg_height = sum(f.get('height', 0) for f in features_list) / len(features_list)
            summary_sheet.append(["Average Width", round(avg_width, 2)])
            summary_sheet.append(["Average Height", round(avg_height, 2)])
        
        # Style summary sheet
        self.worksheet = summary_sheet
        self.style_header()
        self.auto_adjust_column_width()
        
        # Save the workbook
        self.workbook.save(output_path)
        print(f"Excel file with summary saved successfully: {output_path}")
